# -*- coding: utf-8 -*-
"""
IBM HR 员工流失数据集 - 汉化脚本 v5.0
======================================
基于 v4.0 优化，输出可直接用于 Excel 分析的汉化数据集（超级表样式）

功能：
1. 将35个字段名翻译为更符合中文HR术语的命名
2. 将分类变量的值按原数据集定义精准翻译，并保留原始数值作为编码列（适用于数值型分类变量）
3. 输出 Excel 文件（仅包含“数据”工作表，格式化为超级表，命名为 HRDATA）
4. 列顺序按国内阅读习惯及企业系统对接需求排列（员工编号为首列）
5. 应用蓝色主题样式（微软雅黑字体、深蓝标题、行条纹），并自动调整列宽
"""

import pandas as pd
import os
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ==================== 配置区域 ====================
INPUT_FILE = "data/WA_Fn-UseC_-HR-Employee-Attrition.csv"
OUTPUT_DIR = "output"
OUTPUT_EXCEL_FILE = os.path.join(OUTPUT_DIR, "IBM_HR_员工流失数据_本土化版.xlsx")

# ==================== 1. 字段名翻译映射（本土化优化版）====================
COLUMN_TRANSLATION = {
    # 基本信息
    'Age': '年龄',
    'Gender': '性别',
    'MaritalStatus': '婚姻状况',
    'Department': '部门',
    'JobRole': '岗位',
    'JobLevel': '职级',
    
    # 工作相关
    'BusinessTravel': '出差频率',
    'DistanceFromHome': '离家距离',
    'OverTime': '是否加班',
    'StandardHours': '标准工时',
    'JobInvolvement': '敬业度',
    'JobSatisfaction': '工作满意',
    'PerformanceRating': '绩效评级',
    'WorkLifeBalance': '工作与生活平衡',
    
    # 教育背景
    'Education': '教育程度',
    'EducationField': '专业',
    
    # 薪酬福利
    'HourlyRate': '时薪',
    'DailyRate': '日薪',
    'MonthlyRate': '月薪',
    'MonthlyIncome': '月收入',
    'PercentSalaryHike': '调薪幅度',
    'StockOptionLevel': '股权激励等级',
    
    # 工作经历
    'TotalWorkingYears': '总工龄',
    'YearsAtCompany': '本企业工龄',
    'YearsInCurrentRole': '现岗年限',
    'YearsSinceLastPromotion': '晋升间隔',
    'YearsWithCurrManager': '与现任经理共事年限',
    'NumCompaniesWorked': '跳槽次数',
    
    # 满意度评价
    'EnvironmentSatisfaction': '环境满意',
    'RelationshipSatisfaction': '人际关系满意',
    
    # 其他
    'Attrition': '是否离职',
    'EmployeeCount': '员工计数',
    'EmployeeNumber': '员工编号',
    'Over18': '是否成年',
    'TrainingTimesLastYear': '年度培训次数'
}

# ==================== 2. 分类变量值翻译映射 ====================
VALUE_TRANSLATION = {
    # 二元变量
    '是否离职': {'Yes': '是', 'No': '否'},
    '是否加班': {'Yes': '是', 'No': '否'},
    '是否成年': {'Y': '是', 'N': '否'},
    
    # 出差频率
    '出差频率': {
        'Non-Travel': '不出差',
        'Travel_Rarely': '偶尔出差',
        'Travel_Frequently': '频繁出差'
    },
    
    # 部门
    '部门': {
        'Sales': '销售部',
        'Research & Development': '研发部',
        'Human Resources': '人力资源部'
    },
    
    # 教育程度
    '教育程度': {
        1: '大专以下',
        2: '大专',
        3: '本科',
        4: '硕士',
        5: '博士'
    },
    
    # 专业
    '专业': {
        'Life Sciences': '生命科学',
        'Medical': '医学',
        'Marketing': '市场营销',
        'Technical Degree': '工程技术',
        'Human Resources': '人力资源',
        'Other': '其他'
    },
    
    # 环境满意
    '环境满意': {
        1: '低',
        2: '中',
        3: '高',
        4: '非常高'
    },
    
    # 性别
    '性别': {'Male': '男', 'Female': '女'},
    
    # 敬业度
    '敬业度': {
        1: '低',
        2: '中',
        3: '高',
        4: '非常高'
    },
    
    # 岗位
    '岗位': {
        'Sales Executive': '销售主管',
        'Research Scientist': '研究科学家',
        'Laboratory Technician': '实验室技术员',
        'Manufacturing Director': '制造总监',
        'Healthcare Representative': '医疗代表',
        'Manager': '经理',
        'Sales Representative': '销售代表',
        'Research Director': '研究总监',
        'Human Resources': '人力资源专员'
    },
    
    # 工作满意
    '工作满意': {
        1: '低',
        2: '中',
        3: '高',
        4: '非常高'
    },
    
    # 婚姻状况
    '婚姻状况': {
        'Single': '单身',
        'Married': '已婚',
        'Divorced': '离异'
    },
    
    # 绩效评级
    '绩效评级': {
        1: '低',
        2: '良好',
        3: '优秀',
        4: '杰出'
    },
    
    # 人际关系满意
    '人际关系满意': {
        1: '低',
        2: '中',
        3: '高',
        4: '非常高'
    },
    
    # 工作与生活平衡
    '工作与生活平衡': {
        1: '差',
        2: '好',
        3: '更好',
        4: '最好'
    },
    
    # 股权激励等级
    '股权激励等级': {
        0: '无',
        1: '低级',
        2: '中级',
        3: '高级'
    }
}

# ==================== 3. 基础列顺序（员工编号为首列）====================
BASE_COLUMN_ORDER = [
    '员工编号',
    '年龄', '性别', '婚姻状况', '教育程度', '专业', '是否成年',
    '部门', '岗位', '职级', '出差频率', '离家距离', '是否加班', '标准工时', '年度培训次数',
    '总工龄', '本企业工龄', '现岗年限', '晋升间隔', '与现任经理共事年限', '跳槽次数',
    '月收入', '时薪', '日薪', '月薪', '调薪幅度', '股权激励等级',
    '环境满意', '人际关系满意', '工作满意', '敬业度', '工作与生活平衡', '绩效评级',
    '是否离职',
    '员工计数'
]

def reorder_columns(df, base_order):
    """
    根据基础顺序和编码列重新排序列
    - 编码列紧随原列之后
    - 基础顺序中不存在的列放在最后
    """
    final_cols = []
    for col in base_order:
        if col in df.columns:
            final_cols.append(col)
            code_col = col + "编码"
            if code_col in df.columns:
                final_cols.append(code_col)
    for col in df.columns:
        if col not in final_cols:
            final_cols.append(col)
    return df[final_cols]

def apply_excel_formatting(worksheet):
    """
    应用 Excel 格式：超级表、字体、列宽、颜色
    超级表命名为 HRDATA
    """
    # 获取数据区域
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # 创建超级表，命名为 HRDATA
    tab = Table(displayName="HRDATA", ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",  # 蓝色主题
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    worksheet.add_table(tab)
    
    # 设置标题行样式
    for cell in worksheet[1]:
        cell.font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置数据区域样式
    for row in worksheet.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.font = Font(name='微软雅黑', size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽（基于内容长度）
    for col in worksheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        if adjusted_width > 30:
            adjusted_width = 30
        worksheet.column_dimensions[col_letter].width = adjusted_width

def main():
    print("="*60)
    print("IBM HR 员工流失数据集 - 汉化工具 v5.0")
    print("="*60)
    
    # 检查输入文件
    if not os.path.exists(INPUT_FILE):
        print(f"❌ 错误: 找不到输入文件 {INPUT_FILE}")
        print("请确保 data/ 目录下存在原始数据文件")
        return
    
    # 创建输出目录
    Path(OUTPUT_DIR).mkdir(exist_ok=True)
    print(f"📁 输出目录: {OUTPUT_DIR}/")
    
    # 读取数据
    print(f"\n📖 读取数据: {INPUT_FILE}")
    try:
        df_original = pd.read_csv(INPUT_FILE)
        print(f"✅ 读取成功! 共 {len(df_original):,} 行, {len(df_original.columns)} 列")
    except Exception as e:
        print(f"❌ 读取失败: {e}")
        return
    
    # 复制一份用于翻译
    df = df_original.copy()
    
    # 翻译列名
    print("\n🔄 步骤1: 翻译列名...")
    df.rename(columns=COLUMN_TRANSLATION, inplace=True)
    print("✅ 列名翻译完成")
    
    # 翻译变量值，同时为数值型分类变量添加编码列
    print("\n🔄 步骤2: 翻译分类变量值并添加编码列...")
    translated_count = 0
    code_columns_added = []
    for col in df.columns:
        if col in VALUE_TRANSLATION:
            # 检查当前列的数据类型（还未映射，仍是原始值）
            if pd.api.types.is_numeric_dtype(df[col]):
                # 数值型分类变量，先保存编码列
                code_col = col + "编码"
                df[code_col] = df[col]  # 保留原始数值
                code_columns_added.append(code_col)
                # 再进行值映射
                df[col] = df[col].map(VALUE_TRANSLATION[col]).fillna(df[col])
                print(f"  ✓ 翻译列: {col}，并添加编码列 {code_col}")
            else:
                # 非数值型（如字符串），直接映射
                df[col] = df[col].map(VALUE_TRANSLATION[col]).fillna(df[col])
                print(f"  ✓ 翻译列: {col}")
            translated_count += 1
    print(f"✅ 共翻译 {translated_count} 列的分类变量，添加 {len(code_columns_added)} 个编码列")
    
    # 步骤3: 按照国内阅读习惯重新排序列
    print("\n🔄 步骤3: 重新排序列（员工编号为首列，其他按国内习惯）...")
    df = reorder_columns(df, BASE_COLUMN_ORDER)
    print("✅ 列排序完成")
    
    # 保存 Excel 文件并应用格式
    print(f"\n💾 步骤4: 保存并格式化 Excel 文件 - {OUTPUT_EXCEL_FILE}")
    try:
        with pd.ExcelWriter(OUTPUT_EXCEL_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='数据', index=False)
            workbook = writer.book
            worksheet = writer.sheets['数据']
            apply_excel_formatting(worksheet)
        print("✅ Excel 文件保存成功（已应用超级表样式，表格名称：HRDATA）")
    except Exception as e:
        print(f"❌ 保存 Excel 失败: {e}")
        print("请确保已安装 openpyxl: pip install openpyxl")
        return
    
    # 预览
    print("\n📊 数据预览 (前5行，关键列):")
    print("="*80)
    preview_cols = ['员工编号', '年龄', '性别', '部门', '岗位', '是否离职', '月收入', '工作满意', '教育程度', '教育程度编码']
    available_cols = [c for c in preview_cols if c in df.columns]
    print(df[available_cols].head().to_string())
    print("="*80)
    
    # 统计
    if '是否离职' in df.columns:
        attrition_rate = df['是否离职'].value_counts(normalize=True)
        print(f"\n📉 离职率: {attrition_rate.get('是', 0):.2%}")
        print(f"   - 离职人数: {attrition_rate.get('是', 0) * len(df):.0f}")
        print(f"   - 留任人数: {attrition_rate.get('否', 0) * len(df):.0f}")
    
    print(f"\n✨ 完成！输出文件: {OUTPUT_EXCEL_FILE}")
    print("\n📝 版本说明: v5.0")
    print("   - 教育程度: 1 -> 大专以下")
    print("   - 教育领域 → 专业")
    print("   - 输出 Excel 格式（超级表样式：蓝色主题、微软雅黑、自适应列宽，表格名称 HRDATA）")
    print("   - 包含编码列便于分析")
    print("   - 员工编号置于首列，符合企业系统对接习惯")
    print("   - 其他列按国内HR阅读习惯排列")
    print("\n⏭️ 后续将基于此输出开发分析代码（员工画像、流失分析、薪酬公平性、生命周期、职业路径、决策系统等）")

if __name__ == "__main__":
    main()